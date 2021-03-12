from openpyxl import load_workbook
excfile = load_workbook('./test_sheet1.xlsx')

print(excfile.worksheets)
table = excfile.get_active_sheet()
rows = table.rows
print(rows)
columns = table.columns
print(columns)
for row in rows:
    line = [col.value for col in row]
    print(line)

print(table['A1'].value)
Data=table.cell(1,1).value

print(Data)
