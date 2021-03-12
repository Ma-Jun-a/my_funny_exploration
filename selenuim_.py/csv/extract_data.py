import openpyxl

workbook = openpyxl.Workbook()
mysheet = workbook.create_sheet('test',1)
print(workbook.worksheets)
ws = workbook.get_active_sheet()
wc = ws.cell(1,1)
wc.value = 'username'
# wc[1].value = 'password'
workbook.save('test_sheet1.xlsx')

